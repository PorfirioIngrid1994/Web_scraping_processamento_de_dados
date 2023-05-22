from selenium import webdriver
import time
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
import re
import PyPDF2
from datetime import datetime
from openpyxl import Workbook
import sqlite3
from openpyxl import load_workbook

service = Service(executable_path=r"https://dejt.jt.jus.br/dejt/f/n/diariocon")
driver = webdriver.Chrome(service=service)

#Localizando os elementos e preenchendo: 

driver.get("https://dejt.jt.jus.br/dejt/f/n/diariocon")
driver.implicitly_wait(2) # segundos
time.sleep(2)

driver.implicitly_wait(2) # segundos
element = driver.find_element(By.ID, "corpo:formulario:dataIni")
element.clear()  # Limpar o valor atual do campo
element.send_keys("15/05/2023")

time.sleep(2)

driver.implicitly_wait(2) # seconds
element = driver.find_element(By.ID, "corpo:formulario:dataFim")
element.clear()  # Limpar o valor atual do campo
element.send_keys("17/05/2023")

time.sleep(3)

#Preenchendo os campos:

# Elemento pelo <select> pelo ID
element = driver.find_element(By.ID, "corpo:formulario:tipoCaderno")

# Criei um objeto Select a partir do elemento <select>
select = Select(element)

# Opções disponíveis 
options = select.options

# Opção desejada
for option in options:
    if option.text == "Judiciário":
        option.click()
        break

    # Localizei o elemento <select> pelo ID
element = driver.find_element(By.ID, "corpo:formulario:tribunal")

# Criei um objeto Select a partir do elemento <select>
select = Select(element)

# Todas as opções disponíveis 
options = select.options

# Opção desejada por índice
for option in options:
    if option.text == "TST":
        option.click()
        break

wait = WebDriverWait(driver, 10)
element = wait.until(EC.element_to_be_clickable((By.ID, "corpo:formulario:botaoAcaoPesquisar")))
element.click()

time.sleep(2)

#Localizando o botão usando XPath e o onclick
element = driver.find_element(By.XPATH, "//button[contains(@onclick, \"submitForm('corpo:formulario',1,{source:'corpo:formulario:plcLogicaItens:0:j_id131'})\")]")
element.click()
time.sleep(10)

element = driver.find_element(By.XPATH, "//button[contains(@onclick, \"submitForm('corpo:formulario',1,{source:'corpo:formulario:plcLogicaItens:1:j_id131'})\")]")
element.click()
time.sleep(10)

element = driver.find_element(By.XPATH, "//button[contains(@onclick, \"submitForm('corpo:formulario',1,{source:'corpo:formulario:plcLogicaItens:2:j_id131'})\")]")
element.click()
time.sleep(10)

def ler_pdf(arquivo_pdf):
    texto = ""

    with open(arquivo_pdf, 'rb') as file:
        leitor_pdf = PyPDF2.PdfReader(file)
        total_paginas = len(leitor_pdf.pages)

        for pagina in range(total_paginas):
            texto_pagina = leitor_pdf.pages[pagina].extract_text()
            texto += texto_pagina

    return texto

arquivo_pdf = 'C:\\Users\\Ingrid\\Documents\\pdfs_desafio\\Diario_3721__15_5_2023'
texto_completo = ler_pdf(arquivo_pdf)

# Filtragem
padrao = r'\bTST\S*0\b'
palavras_filtradas = re.findall(padrao, texto_completo, re.IGNORECASE)

# Imprimir os dez primeiros resultados
for palavra in palavras_filtradas[:10]:
    print(palavra)

def ler_pdf(arquivo_pdf):
    texto = ""

    with open(arquivo_pdf, 'rb') as file:
        leitor_pdf = PyPDF2.PdfReader(file)
        total_paginas = len(leitor_pdf.pages)

        for pagina in range(total_paginas):
            texto_pagina = leitor_pdf.pages[pagina].extract_text()
            texto += texto_pagina

    return texto

# Lista de arquivos PDF
arquivos_pdf = [
    'C:\\Users\\Ingrid\\Documents\\pdfs_desafio\\Diario_3721__15_5_2023.pdf',
    'C:\\Users\\Ingrid\\Documents\\pdfs_desafio\\Diario_3722__16_5_2023.pdf',
    'C:\\Users\\Ingrid\\Documents\\pdfs_desafio\\Diario_3723__17_5_2023.pdf'
]

# Lista para armazenar os resultados
resultados = []

# Loop para processar os arquivos PDF
for arquivo_pdf in arquivos_pdf:
    texto_completo = ler_pdf(arquivo_pdf)
    
    # Filtragem
    padrao = r'\bTST\S*0\b'
    palavras_filtradas = re.findall(padrao, texto_completo, re.IGNORECASE)

    # Adicionar os resultados à lista
    resultados.extend(palavras_filtradas)

    # Parar o loop se já foram encontrados 10 processos
    if len(resultados) >= 10:
        break

# Imprimir os dez primeiros resultados
for palavra in resultados[:10]:
    print(palavra)

def ler_pdf(arquivo_pdf):
    texto = ""

    with open(arquivo_pdf, 'rb') as file:
        leitor_pdf = PyPDF2.PdfReader(file)
        total_paginas = len(leitor_pdf.pages)

        for pagina in range(total_paginas):
            texto_pagina = leitor_pdf.pages[pagina].extract_text()
            texto += texto_pagina

    return texto

# Lista de arquivos PDF
arquivos_pdf = [
    'C:\\Users\\Ingrid\\Documents\\pdfs_desafio\\Diario_3721__15_5_2023.pdf',
    'C:\\Users\\Ingrid\\Documents\\pdfs_desafio\\Diario_3722__16_5_2023.pdf',
    'C:\\Users\\Ingrid\\Documents\\pdfs_desafio\\Diario_3723__17_5_2023.pdf'
]

# Criar uma nova planilha
planilha = Workbook()
planilha_ativa = planilha.active

# Adicionar cabeçalho à planilha
planilha_ativa.append(['Data e Hora', 'Número do Processo', 'Página', 'Caderno'])

# Loop para processar os arquivos PDF
for arquivo_pdf in arquivos_pdf:
    texto_completo = ler_pdf(arquivo_pdf)
    
    # Aplicar a filtragem utilizando expressões regulares
    padrao = r'\bTST\S*0\b'
    palavras_filtradas = re.findall(padrao, texto_completo, re.IGNORECASE)

    # Inserir os resultados na planilha
    for processo in palavras_filtradas:
        data_hora_registro = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        numero_processo = processo
        pagina = 0  # A ser preenchido com o número correto da página
        caderno = ''  # A ser preenchido com o nome do caderno

        planilha_ativa.append([data_hora_registro, numero_processo, pagina, caderno])

# Salvar a planilha de log
planilha.save('log_processos.xlsx')

print('Dados salvos na planilha de log.')

# Dados extraídos do PDF
dados_extraidos = [
    'TST-AIRR-1160',
    'TST-RR-18177-29.2013.5.16.0020',
    'TST-AIRR-1160',
    'TST-RR-18177-29.2013.5.16.0020',
    'TST-Ag-ARR-208100',
    'TST-E-RR-652000',
    'TST-E-ARR-11012-40',
    'TST-E-RR-1001252-50',
    'TST-RR-11600',
    'TST-RR-20240-81.2020'
]

# Criar uma nova planilha
planilha = Workbook()
planilha_ativa = planilha.active

# Adicionar cabeçalho à planilha
planilha_ativa.append(['Data e Hora', 'Número do Processo', 'Página', 'Caderno'])

# Loop para adicionar os dados na planilha
for processo in dados_extraidos:
    data_hora_registro = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    numero_processo = processo
    pagina = 0  # A ser preenchido com o número correto da página
    caderno = ''  # A ser preenchido com o nome do caderno

    planilha_ativa.append([data_hora_registro, numero_processo, pagina, caderno])

# Salvar a planilha de log
planilha.save('log_processos2.xlsx')

print('Dados salvos na planilha de log.')

# Carregar a planilha de log
planilha = load_workbook('log_processos.xlsx')
planilha_ativa = planilha.active

# Conectando ao banco de dados SQLite (criará um novo banco de dados se não existir)
conexao = sqlite3.connect('dados_processos.db')
cursor = conexao.cursor()

# Criando a tabela se não existir
cursor.execute('''
    CREATE TABLE IF NOT EXISTS processos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        data_hora_envio TEXT,
        data_hora_processamento TEXT,
        numero_processo TEXT,
        pagina INTEGER,
        caderno TEXT
    )
''')

# Dados da planilha e inserção na tabela
for linha in planilha_ativa.iter_rows(min_row=2, values_only=True):
    data_hora_envio = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    data_hora_processamento = linha[0]
    numero_processo = linha[1]
    pagina = linha[2]
    caderno = linha[3]

    cursor.execute('''
        INSERT INTO processos (data_hora_envio, data_hora_processamento, numero_processo, pagina, caderno)
        VALUES (?, ?, ?, ?, ?)
    ''', (data_hora_envio, data_hora_processamento, numero_processo, pagina, caderno))

# Salvando as alterações e fechando a conexão com o banco de dados
conexao.commit()
conexao.close()

print('Dados inseridos na tabela do banco de dados.')